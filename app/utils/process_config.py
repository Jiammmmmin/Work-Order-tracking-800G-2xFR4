"""Persistent per-type per-process configuration: Target UPH, machine count, work hours, and equipment.

Stored in  data/station_config.xlsx  (one sheet per product type).
Sheet columns: Process | Target UPH | Machines | Equipment (comma-separated) | Work Hours
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Set

_DATA_DIR   = Path(__file__).parent.parent.parent / "data"
_EXCEL_PATH = _DATA_DIR / "station_config.xlsx"

# {product_type: {operation: {"target_uph": float|None, "machines": int, "work_hours": float, "equipment": set}}}
_CONFIG: Dict[str, Dict[str, dict]] = {}

_DEFAULT_UPH      = 10.0
_DEFAULT_MACHINES = 1
_DEFAULT_WH       = 8.0


def load() -> None:
    """Read all sheets from the Excel file into the in-memory cache."""
    global _CONFIG
    _CONFIG = {}
    if not _EXCEL_PATH.exists():
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_EXCEL_PATH, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            _CONFIG.setdefault(sheet_name, {})
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                op       = str(row[0]).strip()
                tgt_uph  = float(row[1]) if row[1] not in (None, "") else _DEFAULT_UPH
                machines = int(row[2])   if row[2] not in (None, "") else _DEFAULT_MACHINES
                eq_raw   = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                equipment = {e.strip() for e in eq_raw.split(",") if e.strip()} if eq_raw else set()
                work_hrs = float(row[4]) if len(row) > 4 and row[4] not in (None, "") else _DEFAULT_WH
                if op:
                    _CONFIG[sheet_name][op] = {
                        "target_uph": tgt_uph,
                        "machines":   machines,
                        "equipment":  equipment,
                        "work_hours": work_hrs,
                    }
        wb.close()
    except Exception:
        _CONFIG = {}


def save(product_type: str, ops_data: Dict[str, dict]) -> None:
    """Write ops_data for product_type into the shared Excel file.

    Other sheets (other product types) are preserved.
    """
    _DATA_DIR.mkdir(exist_ok=True)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    if _EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(_EXCEL_PATH)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if product_type in wb.sheetnames:
        del wb[product_type]
    ws = wb.create_sheet(title=product_type)

    headers = ["Process", "Target UPH", "Machines", "Equipment", "Work Hours"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill      = PatternFill("solid", fgColor="1565C0")
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for ri, (op, cfg) in enumerate(ops_data.items(), 2):
        eq_set  = cfg.get("equipment", set())
        eq_str  = ", ".join(sorted(eq_set)) if eq_set else ""
        wh      = cfg.get("work_hours", _DEFAULT_WH)
        ws.cell(row=ri, column=1, value=op)
        ws.cell(row=ri, column=2, value=cfg.get("target_uph", _DEFAULT_UPH))
        ws.cell(row=ri, column=3, value=cfg.get("machines", _DEFAULT_MACHINES))
        ws.cell(row=ri, column=4, value=eq_str)
        ws.cell(row=ri, column=5, value=wh)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 14

    wb.save(_EXCEL_PATH)
    _CONFIG[product_type] = dict(ops_data)


def get(product_type: str, operation: str) -> Optional[dict]:
    """Return config dict or None if not configured."""
    return _CONFIG.get(product_type, {}).get(operation)


def all_ops(product_type: str) -> Dict[str, dict]:
    """Return a copy of all configured operations for this product type."""
    return {op: dict(v) for op, v in _CONFIG.get(product_type, {}).items()}


def merge_operations(
    product_type: str,
    operations: List[str],
    equipment_by_op: Optional[Dict[str, Set[str]]] = None,
) -> None:
    """Add new operations (keeping existing values); merge in any new equipment.

    If equipment data is present for an operation, machines is set to the
    number of distinct equipment IDs.  Without equipment data the existing
    value (or the default of 1) is preserved.
    """
    cfg = _CONFIG.setdefault(product_type, {})
    for op in operations:
        entry = cfg.setdefault(op, {
            "target_uph": _DEFAULT_UPH,
            "machines":   _DEFAULT_MACHINES,
            "work_hours": _DEFAULT_WH,
            "equipment":  set(),
        })
        entry.setdefault("target_uph", _DEFAULT_UPH)
        entry.setdefault("machines",   _DEFAULT_MACHINES)
        entry.setdefault("work_hours", _DEFAULT_WH)
        entry.setdefault("equipment",  set())
        if equipment_by_op and op in equipment_by_op:
            new_eq = {e for e in equipment_by_op[op] if e}
            if new_eq:
                entry["equipment"] |= new_eq
                # Use DB equipment count as the machine count
                entry["machines"] = len(entry["equipment"])
